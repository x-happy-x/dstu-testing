import datetime
import math
import os
import re

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.parser import json

ALL = 'Всё'
ALL_DONE = 'Готово всё'
ALL_FOS = 'Готовы ФОСы'
ALL_TEST = 'Готовы тесты'
ALL_FOS_OR_TEST = 'Готово только ФОСы или тесты'
ANYTHING = 'Готово хоть-нибудь'
NOTHING = 'Нет ничего'


def plan2json(path, sheet, skip=None, save=None, pattern=r'[А-Я]+'):
    if skip is None:
        skip = [0, 0]
    files = os.listdir(path)
    out = {}
    for file in files:

        year = int('20' + file.split('.')[0][::-1].split('-')[0][::-1])
        kurs = int(file.split('_')[-1].split('-')[0])

        name = file.split('_')[0]
        match = re.search(pattern, name)
        if match:
            name = match.group()

        excel = pd.ExcelFile(os.path.join(path, file))

        df = excel.parse(sheet)
        subjs = {}
        for i in range(skip[0], len(df['Закрепленная кафедра']) - skip[1]):
            subj = df.iloc[i, 2]
            block = df.iloc[i, 1]
            kaf = df['Закрепленная кафедра'][i]

            if str(subj) == 'nan' or str(kaf) == 'nan':
                continue

            add_data = {"Кафедра": int(kaf), "Блок": block}
            if subj in subjs:
                b = subjs.pop(subj)
                b['R'] = subj
                add_data['R'] = subj
                subjs[f"{subj} ({b['Блок'].split('(')[-1].split(')')[0]})"] = b
                subjs[f"{subj} ({block.split('(')[-1].split(')')[0]})"] = add_data
                # print(file, subj, add_data, b, "повтор")
            else:
                subjs[subj] = add_data

        if name not in out:
            out[name] = {}
        if year not in out[name]:
            out[name][year] = {}
        if kurs not in out[name][year]:
            out[name][year][kurs] = {}
        out[name][year][kurs][file] = subjs

    if save is not None:
        json.to_file(out, save)

    return out


__parse_data = {
    r'^(B090302_|B090302ВИС_)': 'ВИС',
    r'^B090302ВИАС_': 'ВИАС',
    r'^B090303ВЗПИ_': 'ВЗПИ',
    r'^B090303ВОЗПИ_': 'ВОЗПИ',
    r'^090402МЗИН_': 'МЗИН',
    r'^090402МИН_': 'МИН',
    r'^090403МПИ_': 'МПИ',
}


def parse_name(filename):
    year = int('20' + filename.split('.')[0][::-1].split('-')[0][::-1])
    course = int(filename.split('_')[-1].split('-')[0])

    name = filename.split('_')[0]
    for ptr in __parse_data:
        match = re.search(ptr, filename.strip())
        if match:
            name = __parse_data[ptr]

    return name, year, course


def preps2json(file, sheet, skip=None, save=None):
    if skip is None:
        skip = [0, 0]

    excel = pd.ExcelFile(file)
    df = excel.parse(sheet)

    pr = {}

    last_name = None
    last_id = None
    last_index = 0
    subjects = {}

    for i in range(skip[0], len(df['№']) - skip[1]):
        prep = df.iloc[i]
        if 'nan' != str(prep['Фамилия']):
            last_name = f"{prep['Фамилия'].strip()} {prep['Имя'].strip()} {prep['Отчество'].strip()}"
            if last_id is None:
                last_index = i
                subjects[last_index] = []
        if 'nan' != str(prep['№']):
            last_id = prep['№'].strip().strip('.')
            if last_name is None:
                last_index = i
                subjects[last_index] = []
        subjects[last_index].append(prep['Дисциплина'])
        if last_id is None or last_name is None:
            continue
        pr[last_index] = {'id': last_id, 'name': last_name}
        last_id = None
        last_name = None

    for i in subjects:
        pr[i]['subjs'] = subjects[i]

    preps = {}
    for i in pr:
        preps[pr[i]['name']] = {}
        for s in pr[i]['subjs']:
            if str(s) == 'nan':
                continue
            ss = s.split('[')
            preps[pr[i]['name']][ss[0].strip()] = ss[1].strip().strip(']').strip().upper().split(',') if len(
                ss) > 1 else []

    if save is not None:
        json.to_file(preps, save)

    return preps


def get_prep(preps, subj, group, short=True):
    subj = subj.strip()  # .lower()
    group = group.strip().upper()
    finded = []
    for prep in preps:
        if subj in preps[prep] and (len(preps[prep][subj]) == 0 or group in preps[prep][subj]):
            if short:
                f, i, o = prep.split(' ')
                finded.append(f"{f} {i[0]}.{o[0]}.")
                continue
            finded.append(prep)
    return None if len(finded) == 0 else ", ".join(finded)


def get_departament_name(departments, code):
    return list(filter(lambda x: x['number'] == code, departments))[0]['name']


def plan2excel(plans, preps, departments, reports, kaf, save=None):
    if save is None:
        save = "plan.xlsx"
    wb = load_workbook('./template/LayoutPlan.xlsx')
    wb: Workbook
    # writer = pd.ExcelWriter(f'{save}.xlsx', engine='xlsxwriter')
    pm = ['Нет', 'Да']
    columns = [
        '№',
        'Наименование дисциплины',
        'Ответственный за разработку комплекта ТЗ',
        'ОМ согласованы (наличие согласовательных подписей, № протоколов, даты)',
        'Размещены в приложениях к РПД',
        'Размещен на сайте СКИФ.ТЕСТ (верифицирован)',
        'Размещены в приложениях к РПД (Тесты)',
        'Кафедра',
        'Ссылки',
        'Ссылки в РПД на тест',
        'Компетенции',
        'Правильность часов',
        'Тип контроля',
        'Файлы'
    ]
    for p_group in plans:
        for p_year in plans[p_group]:
            for p_course in plans[p_group][p_year]:

                i = 1
                map_subjs = []

                for p_filename in plans[p_group][p_year][p_course]:
                    plan = plans[p_group][p_year][p_course][p_filename]

                    for subject in plan:
                        # if kaf != plan[subject]['Кафедра']:
                        #     continue
                        academic_year = f"{p_year}-{int(p_year) + 1}"
                        p_kaf_name = get_departament_name(departments, plan[subject]['Кафедра'])
                        real_subject = plan[subject]['R'] if 'R' in plan[subject] else subject

                        if academic_year not in reports:
                            print(f"Год: {academic_year} не найден в report")
                            continue
                        # Выбор по году
                        reports_year = reports[academic_year]
                        _kafs = _filter(reports_year, lambda r: r['name'] == p_kaf_name)
                        if len(_kafs) == 0:
                            print(f"Кафедра: {p_kaf_name} не найден в report[{academic_year}]")
                            continue
                        # Выбор по кафедре
                        for _kaf_id, reports_kaf in _kafs:
                            subject_name_clear = lambda x: x.replace(',', '').replace(' ', '')
                            rs = subject_name_clear(real_subject)
                            _subjects = _filter(
                                reports_kaf['items'],
                                lambda r: subject_name_clear(r['name']) == rs
                            )
                            if len(_subjects) == 0:
                                print(f"Дисциплина: {real_subject} "
                                      f"не найдена в report[{academic_year}][{_kaf_id} | {reports_kaf['name']}]")
                                continue
                            # Выбор по дисциплине
                            for _subject_id, reports_subject in _subjects:
                                _plans = _filter(reports_subject['items'])
                                if len(_plans) == 0:
                                    print(f"План не найден в "
                                          f"report[{academic_year}][{_kaf_id} | {reports_kaf['name']}][{_subject_id}]")
                                    continue
                                fos = None
                                skif = None
                                url = None
                                skif_url = None
                                files = []
                                comps = None
                                summary = None
                                control = None
                                # Планы
                                for _plan_id, reports_plan in _plans:
                                    _rpd = _filter(
                                        reports_plan['items'],
                                        lambda r: r['group']['name'] == p_group
                                    )
                                    if len(_plans) == 0:
                                        print(f"РПД не найдено в "
                                              f"report[{academic_year}][{_kaf_id} | {reports_kaf['name']}]"
                                              f"[{_subject_id}][{_plan_id} | {reports_plan['name']}]")
                                        continue
                                    _, year, course = parse_name(reports_plan['name'])
                                    if str(year) != str(p_year) or str(course) != str(p_course):
                                        continue
                                    # РПД
                                    for _rpd_id, report_rpd in _rpd:
                                        if url is None:
                                            url = ''
                                        url += report_rpd['url'] + '\n'
                                        if skif_url is None:
                                            skif_url = ''
                                        skif_url += str(report_rpd['ФОС'][0]['url']) + '\n'
                                        if skif is None:
                                            skif = ''
                                        skif += pm[report_rpd['ФОС'][0]['exist']] + '\n'
                                        if comps is None:
                                            comps = ''
                                        comps += report_rpd['Компетенции'][0] + '\n'
                                        if summary is None:
                                            summary = ''
                                        summary += pm[report_rpd['Часы'][0]['Правильность часов:']] + '\n'
                                        if control is None:
                                            control = ''
                                        for sr in report_rpd['Часы'][0]['Контроль:']:
                                            control += f"{sr['Тип контроля']} ({sr['Семестр']} сем.)\n"
                                        if fos is None:
                                            fos = ''
                                        fos += pm[report_rpd['Приложение'][0]['new']] + '\n'
                                        files.extend(report_rpd['Приложение'][1]['files'])
                                map_subjs.append([
                                    i, subject, get_prep(preps, real_subject, p_group),
                                    None, str(fos).strip(), str(skif).strip(), str(skif).strip(), p_kaf_name.strip(),
                                    str(url).strip(), str(skif_url).strip(), str(comps).strip(), str(summary).strip(),
                                    str(control).strip(), files
                                ])
                                i += 1
                sheet_name = f"{p_group} ({p_course}-{str(p_year)[2:]})"
                if sheet_name not in wb.sheetnames:
                    ws = wb.copy_worksheet(wb['ШАБЛОН'])
                    ws.title = sheet_name

                ws = wb[sheet_name]
                ws: Worksheet

                for col in range(len(columns)):
                    ws.cell(1, col + 1).value = columns[col]

                centered = [0, 2, 3, 4, 5, 6, 10, 11, 12]
                links = [8, 9]
                more = [13]
                for row in range(len(map_subjs)):
                    offset = 1
                    for col in range(len(map_subjs[row])):
                        if map_subjs[row][col] is None or str(map_subjs[row][col]) == 'None':
                            continue
                        if col in more:
                            for f in map_subjs[row][col]:
                                ws.cell(row + 2, col + offset).value = f['url'] + '\n' + f['name'] + '\n' + f[
                                    'modified_site'] + '\n' + f['modified_local'] + '\n' + f['path']
                                ws.cell(row + 2, col + offset).alignment = Alignment(horizontal='left', wrap_text=True,
                                                                                     vertical='center')
                                ws.cell(row + 2, col + offset).hyperlink = f['url']
                                offset += 1
                        else:
                            ws.cell(row + 2, col + offset).value = map_subjs[row][col]
                        if col in centered:
                            ws.cell(row + 2, col + offset).alignment = Alignment(horizontal='center', wrap_text=True,
                                                                                 vertical='center')
                        else:
                            ws.cell(row + 2, col + offset).alignment = Alignment(horizontal='left', wrap_text=True,
                                                                                 vertical='center')
                        if col in links:
                            ws.cell(row + 2, col + offset).hyperlink = str(map_subjs[row][col]).split('\n')[0]

    wb['ШАБЛОН'].sheet_state = 'hidden'
    wb.save(save)


def _filter(items, function=None):
    if function is None:
        return items.items()
    return list(filter(lambda item: function(item[1]), items.items()))


def plan2stat(plans, preps, departments, reports, kaf, save=None):
    report = {}
    for p_group in plans:
        group_report = {}
        for p_year in plans[p_group]:
            year_report = {}
            for p_course in plans[p_group][p_year]:
                report_data = {
                    ALL: 0,
                    ALL_DONE: 0,
                    ALL_FOS: 0,
                    ALL_TEST: 0,
                    ALL_FOS_OR_TEST: 0,
                    ANYTHING: 0,
                    NOTHING: 0,
                }
                report_data['stat'] = report_data.copy()
                course_report = report_data.copy()
                course_report['наши'] = report_data.copy()
                course_report['чужие'] = report_data.copy()
                for p_filename in plans[p_group][p_year][p_course]:
                    plan = plans[p_group][p_year][p_course][p_filename]
                    for subject in plan:
                        # if kaf != plan[subject]['Кафедра']:
                        #     continue
                        academic_year = f"{p_year}-{int(p_year) + 1}"
                        p_kaf_name = get_departament_name(departments, plan[subject]['Кафедра'])
                        real_subject = plan[subject]['R'] if 'R' in plan[subject] else subject

                        if academic_year not in reports:
                            print(f"Год: {academic_year} не найден в report")
                            continue
                        # Выбор по году
                        reports_year = reports[academic_year]
                        _kafs = _filter(reports_year, lambda r: r['name'] == p_kaf_name)
                        if len(_kafs) == 0:
                            print(f"Кафедра: {p_kaf_name} не найден в report[{academic_year}]")
                            continue
                        # Выбор по кафедре
                        for _kaf_id, reports_kaf in _kafs:
                            subject_name_clear = lambda x: x.replace(',', '').replace(' ', '')
                            rs = subject_name_clear(real_subject)
                            _subjects = _filter(
                                reports_kaf['items'],
                                lambda r: subject_name_clear(r['name']) == rs
                            )
                            if len(_subjects) == 0:
                                print(f"Дисциплина: {real_subject} "
                                      f"не найдена в report[{academic_year}][{_kaf_id} | {reports_kaf['name']}]")
                                continue
                            # Выбор по дисциплине
                            for _subject_id, reports_subject in _subjects:
                                _plans = _filter(reports_subject['items'])
                                if len(_plans) == 0:
                                    print(f"План не найден в "
                                          f"report[{academic_year}][{_kaf_id} | {reports_kaf['name']}][{_subject_id}]")
                                    continue
                                # Планы
                                for _plan_id, reports_plan in _plans:
                                    _rpd = _filter(
                                        reports_plan['items'],
                                        lambda r: r['group']['name'] == p_group
                                    )
                                    if len(_plans) == 0:
                                        print(f"РПД не найдено в "
                                              f"report[{academic_year}][{_kaf_id} | {reports_kaf['name']}]"
                                              f"[{_subject_id}][{_plan_id} | {reports_plan['name']}]")
                                        continue
                                    _, year, course = parse_name(reports_plan['name'])
                                    if str(year) != str(p_year) or str(course) != str(p_course):
                                        continue
                                    # РПД
                                    for _rpd_id, report_rpd in _rpd:
                                        fos = report_rpd['Приложение'][0]['new']
                                        test = report_rpd['ФОС'][0]['exist']

                                        course_report[ALL] += 1
                                        course_report[ALL_DONE] += int(fos and test)
                                        course_report[ALL_FOS] += int(fos and not test)
                                        course_report[ALL_TEST] += int(test and not fos)
                                        course_report[ALL_FOS_OR_TEST] += int(not fos and test or not test and fos)
                                        course_report[ANYTHING] += int(fos or test)
                                        course_report[NOTHING] += int(not fos and not test)

                                        if kaf != plan[subject]['Кафедра']:
                                            course_report['наши'][ALL] += 1
                                            course_report['наши'][ALL_DONE] += int(fos and test)
                                            course_report['наши'][ALL_FOS] += int(fos)
                                            course_report['наши'][ALL_TEST] += int(test)
                                            course_report['наши'][ALL_FOS_OR_TEST] += int(not fos or not test)
                                            course_report['наши'][ANYTHING] += int(fos or test)
                                            course_report['наши'][NOTHING] += int(not fos and not test)
                                        else:
                                            course_report['чужие'][ALL] += 1
                                            course_report['чужие'][ALL_DONE] += int(fos and test)
                                            course_report['чужие'][ALL_FOS] += int(fos)
                                            course_report['чужие'][ALL_TEST] += int(test)
                                            course_report['чужие'][ALL_FOS_OR_TEST] += int(not fos or not test)
                                            course_report['чужие'][ANYTHING] += int(fos or test)
                                            course_report['чужие'][NOTHING] += int(not fos and not test)


                year_report[p_course] = course_report
            group_report[p_year] = year_report
        report[p_group] = group_report
    if save:
        json.to_file(report, save)
    return report


def isnan(x):
    try:
        return math.isnan(float(x))
    except ValueError:
        return False


def get_type(competencies, index):
    if not isnan(competencies['Индекс'][index]):
        return 1, competencies['Индекс'][index]
    if not isnan(competencies['Unnamed: 1'][index]):
        return 2, competencies['Unnamed: 1'][index]
    if not isnan(competencies['Unnamed: 2'][index]):
        return 3, competencies['Unnamed: 2'][index]
    return 4, ""


def get_competencies(page: pd.DataFrame):
    competencies = {}
    last_type = {}
    for i in range(0, len(page)):
        level, title = get_type(page, i)
        text = page['Содержание'][i]
        if level == 1:
            last_type['type'] = title.split('-')[0].replace('_x000d_', '').strip(' \n;.\t,')
            last_type['title1'] = title.replace('_x000d_', '').strip(' \n;.\t,')
            if last_type['type'] not in competencies:
                competencies[last_type['type']] = {}
            competencies[last_type['type']][last_type['title1']] = {
                'text': text.replace('_x000d_', '').strip(' \n;.\t,'),
                'sub': {}
            }
        elif level == 2:
            last_type['title2'] = title.replace('_x000d_', '').strip(' \n;.\t,')
            competencies[last_type['type']][last_type['title1']]['sub'][last_type['title2']] = {
                'text': text.replace('_x000d_', '').strip(' \n;.\t,'),
                'sub': []
            }
        else:
            competencies[last_type['type']][last_type['title1']]['sub'][last_type['title2']]['sub'].append(
                text.replace('_x000d_', '').strip(' \n;.\t,'))
    return competencies


def get_info(page):
    study_form = page['Unnamed: 2'][40].split(':')[-1].strip().split(' ')[0]
    study_code, study_name = page['Unnamed: 3'][27].strip().split(' ', 1)
    study_direction = page['Unnamed: 3'][28]
    study_code = study_code.replace('.', '')
    return {
        'code': study_code.replace('_x000d_', '').strip(' \n;.\t,'),
        'name': study_name.replace('_x000d_', '').strip(' \n;.\t,'),
        'form': study_form.replace('_x000d_', '').strip(' \n;.\t,'),
        'direction': study_direction.replace('_x000d_', '').strip(' \n;.\t,'),
        'year': str(page['Unnamed: 22'][38]).replace('_x000d_', '').strip(' \n;.\t,'),
    }


def get_name(group, info):
    year = int(info['year'])
    current_year = datetime.datetime.now().year
    return f"__ИД_{group}_{current_year - year + 1}к__23-24.docx"


def load_plan(group, path):
    excel = pd.ExcelFile(path)

    info = get_info(excel.parse('Титул'))
    competencies = get_competencies(excel.parse('Компетенции'))
    filename = get_name(group, info)

    return {
        'filename': filename,
        'info': info,
        'competencies': competencies,
    }

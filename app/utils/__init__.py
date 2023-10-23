from datetime import datetime

from openpyxl.styles import Alignment, Font


def __diff_stat__(source_plan, dest_plan, stat, add=True):
    for comp in source_plan['competencies']:
        if comp in dest_plan['competencies']:
            competencies = source_plan['competencies'][comp]
            for competency in competencies:
                if competency in dest_plan['competencies'][comp]:
                    if add and competencies[competency]['text'] != dest_plan['competencies'][comp][competency]['text']:
                        stat['Компетенции']['.'][competency] = {
                            'new': competencies[competency]['text'],
                            'old': dest_plan['competencies'][comp][competency]['text']
                        }
                    indicators = competencies[competency]['sub']
                    for indicator in indicators:
                        if indicator in dest_plan['competencies'][comp][competency]['sub']:

                            if add and indicators[indicator]['text'] != \
                                    dest_plan['competencies'][comp][competency]['sub'][indicator]['text']:
                                stat['Индикаторы']['.'][indicator] = {
                                    'new': indicators[indicator]['text'],
                                    'old': dest_plan['competencies'][comp][competency]['sub'][indicator]['text']
                                }
                            disciplines = indicators[indicator]['sub']
                            for discipline in disciplines:
                                if discipline not in dest_plan['competencies'][comp][competency]['sub'][indicator][
                                    'sub']:
                                    stat['Дисциплины']['+' if add else '-'].append({
                                        'дисциплина': discipline,
                                        'индикатор': indicator
                                    })
                        else:
                            stat['Индикаторы']['+' if add else '-'][indicator] = indicators[indicator]['text']
                else:
                    stat['Компетенции']['+' if add else '-'][competency] = competencies[competency]['text']


def diff_stat(source_plan, dest_plan):
    if source_plan == dest_plan:
        return None

    stat = {
        'info': {
            'source': source_plan['filename'],
            'dest': dest_plan['filename'],
        },
        'Компетенции': {
            '+': {},
            '-': {},
            '.': {},
        }, 'Индикаторы': {
            '+': {},
            '-': {},
            '.': {},
        }, 'Дисциплины': {
            '+': [],
            '-': [],
        }
    }

    __diff_stat__(source_plan, dest_plan, stat, True)
    __diff_stat__(dest_plan, source_plan, stat, False)

    return stat


def discipline_filter(disciplines: list[str], shorten: list[str] = None, blacklist: list[str] = None):
    new_list = []
    if shorten is None:
        shorten = []
    if blacklist is None:
        blacklist = []
    for discipline in disciplines:
        skip = False
        for black_disp in blacklist:
            if discipline.lower() == black_disp.lower():
                skip = True
                break
        if skip:
            continue
        for shorten_disp in shorten:
            if discipline.lower().startswith(shorten_disp.lower()):
                if shorten_disp not in new_list:
                    new_list.append(shorten_disp)
                skip = True
                break
        if skip:
            continue
        if discipline not in new_list:
            new_list.append(discipline)
    return new_list


def get_temp_filename(extension='docx'):
    return f'./temp-{datetime.now():%Y_%m_%d %H_%M_%S%z}.{extension}'


def format_sheet(ws):
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 70
    state = 0
    header_skip = False
    for row in ws.rows:
        row_idx = row[0].row
        row[0].alignment = Alignment(horizontal='center', vertical='center')
        row[1].alignment = Alignment(horizontal='center', vertical='center')
        if (state == 0 and row[0].value == "Компетенции" or
                state == 1 and row[0].value == "Индикаторы" or
                state == 2 and row[0].value == "Дисциплины"):
            row[0].font = Font(bold=True)
            for i in range(4):
                ws.cell(row_idx + 1, i + 1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row_idx + 1, i + 1).font = Font(italic=True)
            state += 1
            header_skip = True
            continue
        elif ws.cell(row_idx, 1).value == "Нет изменений":
            ws.merge_cells("A" + str(row_idx) + ":D" + str(row_idx))
            ws.cell(row_idx, 1).font = Font(underline='single')
            ws.cell(row_idx, 1).alignment = Alignment(horizontal='center', vertical='center')
        elif state == 2 and not header_skip:
            for i in range(2, 4):
                ws.cell(row_idx, i + 1).alignment = Alignment(wrap_text=True, vertical='center')
        header_skip = False

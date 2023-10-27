import json
import os
import time
from datetime import datetime
import random

from typing import Callable

import pandas as pd

from app.parser.json import save_json, open_json
from app.rpd.RpdApp import Department, Discipline, Plan, RP, Appx, get_now_year, Params, CompetenceBoard, Result
from app.rpd import api, RpdApp

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

testing_mode = False

if testing_mode:
    rpdapi = api.RpdApi(data={
        Params.rup_row_id: 3019716
    })
    response = rpdapi.session.get("https://rpd.donstu.ru/Rp/Initialize?rupRowId=3019716")
    if response.status_code == 200:
        data = json.loads(response.text)
        response = rpdapi.session.post("https://rpd.donstu.ru/Title/AddAuthor?rupRowId=3019716", data={
            'rank': '1',
            'position': '2',
            'fio': '3',
            'hash': data['data']['commonState']['hash'],
        })
        if response.status_code == 200:
            print(response.text)
            # with open('file.txt', 'r') as f:
            #     s.cookies['.AspNetCore.Cookies'] = ''
            #
            # with open('file.txt', 'w') as f:
            #     f.write(s.cookies['.AspNetCore.Cookies'])
        else:
            print(response.text)
    else:
        print(response.text)
    exit()


def department_selector(d: Department) -> bool:
    return d.id == 9


def discipline_selector(d: Discipline) -> bool:
    return True


def plan_selector(plan: Plan) -> bool:
    groups = ['B090302_', 'B090302ВИАС_', 'B090302ВИС_', 'B090303ВЗПИ_', 'B090303ВОЗПИ_', '090402МЗИН_', '090402МИН_',
              '090403МПИ_']
    for group in groups:
        if group in plan.rup_name:
            return True
    return False


def rpd_selector(rp: RP) -> bool:
    return True


def file_selector(f: Appx) -> bool:
    return 'ФОС' in f.name


department: Department
discipline: Discipline
plan: Plan
rpd: RP
file: Appx
cb: CompetenceBoard | None = None

app = RpdApp()

# Параметры обхода
PAUSE_TIME = 0.0
# department_selector = lambda x: True
# discipline_selector = lambda x: True
plan_selector = lambda x: True
# rpd_selector = lambda x: True
# file_selector = lambda x: True

check_summary = True
wb_summary = Workbook()

# Проверка ФОСов
check_fos = False
REPORT = {}

# Выгрузка компетенций
competence_settings = {
    'Проверять': True,
    'Исходный файл': None,
    'Лист': "Все компетенции",
    'Сохранить': 'comps2.xlsx',
    'Строка': 1,
}

# Создание таблиц по практикам
check_practice_tables = True

if competence_settings['Исходный файл'] is None:
    competence_settings['Файл'] = Workbook()
    del competence_settings['Файл']['Sheet']
else:
    competence_settings['Файл'] = load_workbook(competence_settings['Исходный файл'])
if competence_settings['Лист'] not in competence_settings['Файл'].sheetnames:
    competence_settings['Файл'].create_sheet(competence_settings['Лист'])
competence_settings['Рабочий лист'] = competence_settings['Файл'][competence_settings['Лист']]

# Учебный год
for year in [get_now_year()]:
    print("Год", year)
    REPORT_YEAR = {}
    departments = app.departments(year)
    time.sleep(random.random() * PAUSE_TIME)  # Пауза до {PAUSE_TIME} сек

    # Кафедра
    for department in departments(department_selector):
        print("\tКафедра", department.name)
        REPORT_DEPARTMENT = {}
        disciplines = department.disciplines()
        time.sleep(random.random() * PAUSE_TIME)  # Пауза до {PAUSE_TIME} сек

        # Дисциплина (модуль)
        for discipline in disciplines(discipline_selector):
            # if not discipline.name.startswith("М"):
            #     continue
            print("\t\tДисциплина", discipline.name)
            REPORT_DISCIPLINE = {}
            plans = discipline.plans()
            time.sleep(random.random() * PAUSE_TIME)  # Пауза до {PAUSE_TIME} сек

            # Учебный план
            for plan in plans(plan_selector):
                print("\t\t\tУчебный план", plan.rup_name)
                REPORT_PLAN = {}
                rps = plan.rps()
                time.sleep(random.random() * PAUSE_TIME)  # Пауза до {PAUSE_TIME} сек

                # Рабочая программа
                for rpd in rps(rpd_selector):
                    print(f"\t\t\t\tРабочая программа {rpd.id}:", rpd.name)

                    REPORT_RPD = None

                    if check_summary:
                        sheet_summary = wb_summary.create_sheet(f"{rpd.id}")
                        # app.load_cache = False
                        rss = rpd.summary()
                        summary = rss.data
                        print()
                        df = pd.json_normalize(summary['data']['items'])
                        df = df.set_index('ВидЗанятий')
                        print(df)
                        for i in range(0, len(df.columns), 2):
                            diff_summary = df.iloc[:, i + 1] - df.iloc[:, i]
                            diff_summary_b = diff_summary != 0
                            # for j in range(0, len(diff_summary)):
                            #     if diff_summary_b.iloc[j]:
                            #         name_s = diff_summary.keys()[j]
                            #         value_s = diff_summary.iloc[j]
                            #         if value_s < 0:
                            #             text = f"В тип '{name_s}' нужно добавить {abs(value_s)} часов"
                            #         else:
                            #             text = f"В типе '{name_s}' нужно убрать {abs(value_s)} часов"
                            #         print(text)
                        # sheet_summary.sheet_properties.tabColor = 'ffff0000'

                    if competence_settings['Проверять']:
                        competence_settings['Рабочий лист'][f"A{competence_settings['Строка']}"].value = rpd.name
                        competence_settings['Строка'] += 1

                        cbr = rpd.competencies_board()
                        cb = cbr.data
                        for comp in cb.competencies():
                            competence_settings['Рабочий лист'][
                                f"A{competence_settings['Строка']}"].value = comp.comp_code
                            competence_settings['Рабочий лист'][
                                f"B{competence_settings['Строка']}"].value = comp.comp_description
                            competence_settings['Строка'] += 1
                            for indicator in comp.indicators():
                                competence_settings['Рабочий лист'][
                                    f"A{competence_settings['Строка']}"].value = indicator.indi_code
                                competence_settings['Рабочий лист'][
                                    f"B{competence_settings['Строка']}"].value = indicator.indi_description
                                competence_settings['Строка'] += 1
                                for level in indicator.levels():
                                    competence_settings['Рабочий лист'][
                                        f"A{competence_settings['Строка']}"].value = level.level_id
                                    competence_settings['Рабочий лист'][
                                        f"B{competence_settings['Строка']}"].value = level.contents
                                    competence_settings['Строка'] += 1

                        competence_settings['Строка'] += 2

                    if check_practice_tables and "практика" in discipline.name.lower():
                        if cb is None:
                            cbr = rpd.competencies_board()
                            cb = cbr.data
                        cb.generate_table1(rpd.get_file_path('table1.docx')[1])
                        cb.generate_table2(rpd.get_file_path('table2.docx')[1])

                    if check_fos:
                        # Приложение
                        files = rpd.appxs()
                        fos = files(file_selector)

                        # Проверка состояния ФОСа
                        new = False
                        loaded = False
                        fn = None
                        if len(fos) == 1:
                            file = fos[0]
                            fn = str(file)
                            if file.modified is not None:
                                new = file.modified > datetime(2023, 1, 1)
                                if new:
                                    REPORT_RPD = f"Фос: {file.name} новый (дата в РПД: {file.modified})"
                                else:
                                    REPORT_RPD = f"Фос: {file.name} старый (дата в РПД: {file.modified})"
                            else:
                                if not file.exists():
                                    file.save()
                                loaded = True
                                prop = file.metadata()
                                if isinstance(prop.modified, datetime):
                                    new = prop.modified > datetime(2023, 1, 1)
                                else:
                                    print(prop.modified)

                                if new:
                                    REPORT_RPD = f"Фос: {file.name} новый (дата в метаданных: {prop.modified})"
                                else:
                                    REPORT_RPD = f"Фос: {file.name} старый (дата в метаданных: {prop.modified})"
                        elif len(fos) == 0:
                            REPORT_RPD = "Нет фоса"
                        else:
                            fn = str(files.data)
                            REPORT_RPD = "Нужно проверить вручную"

                        REPORT_PLAN[rpd.id] = {"Состояние ФОСа": REPORT_RPD,
                                               "filename": fn,
                                               "download": loaded,
                                               "new": new,
                                               "test": rpd.FOS()["Ссылка на СКИФ.Тест"],
                                               "url": f"https://rpd.donstu.ru/Rp?rupRowId={rpd.rup_row_id}&rpId={rpd.id}"}

                REPORT_DISCIPLINE[plan.rup_name] = REPORT_PLAN
            REPORT_DEPARTMENT[discipline.name] = REPORT_DISCIPLINE
        REPORT_YEAR[department.name] = REPORT_DEPARTMENT
    REPORT[year] = REPORT_YEAR



if competence_settings['Проверять']:
    competence_settings['Файл'].save(competence_settings['Сохранить'])
wb_summary.save('ttt.xlsx')
if check_fos:
    save_json(REPORT, './fos_status.json')

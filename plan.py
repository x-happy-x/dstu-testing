import re
from copy import deepcopy

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.parser import json
from app import parser
import app.parser.plan.parser as ps
import pandas as pd

# plans = parser.plan2json('./УП', 'План', [2, 0], save='./plan.json')
# preps = parser.preps2json("./Учет.xlsx", "Учет", [2, 0], './preps.json')

plans = json.from_file('./source/plan.json')
preps = json.from_file('./source/preps.json')
departments = json.from_file('./.cache/RPD_API/2023-2024/departments.json')
fos_status = json.from_file('./source/reports.json')

# parser.plan2excel(
#     plans,
#     preps,
#     departments['data']['department']['items'],
#     fos_status,
#     9,
#     "./dest/Карта учета ОМ и тестов.xlsx"
# )

# parser.plan2stat(
#     plans,
#     preps,
#     departments['data']['department']['items'],
#     fos_status,
#     9,
#     "./dest/стат.json"
# )

wb = load_workbook("./dest/Карта учета ОМ и тестов.xlsx")
wb: Workbook
sheet: Worksheet
stat = {}

kk = ['Чужая', 'Наша']
report_data = {
    ps.ALL: {'Все': 0, 'Чужая': 0, 'Наша': 0},
    ps.ALL_DONE: {'Все': 0, 'Чужая': 0, 'Наша': 0},
    ps.ALL_FOS: {'Все': 0, 'Чужая': 0, 'Наша': 0},
    ps.ALL_TEST: {'Все': 0, 'Чужая': 0, 'Наша': 0},
    ps.ALL_FOS_OR_TEST: {'Все': 0, 'Чужая': 0, 'Наша': 0},
    ps.ANYTHING: {'Все': 0, 'Чужая': 0, 'Наша': 0},
    ps.NOTHING: {'Все': 0, 'Чужая': 0, 'Наша': 0},
}

for sheet in wb:
    if sheet.title == 'ШАБЛОН':
        continue
    print(sheet.title)
    stat[sheet.title] = deepcopy(report_data)
    match = re.search(r'\d+$', sheet.dimensions)
    rows = int(match.group(0))
    for i in range(2, rows):

        f_cell = sheet[f"E{i}"]
        fos = False
        if f_cell.value is not None and '\n' in f_cell.value:
            if 'да' in f_cell.value.lower():
                fos = True
            else:
                fos = False
        else:
            fos = f_cell.value is not None and f_cell.value.lower() == 'да'

        t_cell = sheet[f"F{i}"]
        test = False
        if t_cell.value is not None and '\n' in t_cell.value:
            if 'да' in t_cell.value.lower():
                test = True
            else:
                test = False
        else:
            test = t_cell.value is not None and t_cell.value.lower() == 'да'

        kaf = sheet[f"H{i}"].value == 'Информационные технологии'

        stat[sheet.title][ps.ALL]['Все'] += 1
        stat[sheet.title][ps.ALL_DONE]['Все'] += int(fos and test)
        stat[sheet.title][ps.ALL_FOS]['Все'] += int(fos and not test)
        stat[sheet.title][ps.ALL_TEST]['Все'] += int(test and not fos)
        stat[sheet.title][ps.ALL_FOS_OR_TEST]['Все'] += int(not fos and test or not test and fos)
        stat[sheet.title][ps.ANYTHING]['Все'] += int(fos or test)
        stat[sheet.title][ps.NOTHING]['Все'] += int(not fos and not test)

        stat[sheet.title][ps.ALL][kk[kaf]] += 1
        stat[sheet.title][ps.ALL_DONE][kk[kaf]] += int(fos and test)
        stat[sheet.title][ps.ALL_FOS][kk[kaf]] += int(fos and not test)
        stat[sheet.title][ps.ALL_TEST][kk[kaf]] += int(test and not fos)
        stat[sheet.title][ps.ALL_FOS_OR_TEST][kk[kaf]] += int(not fos and test or not test and fos)
        stat[sheet.title][ps.ANYTHING][kk[kaf]] += int(fos or test)
        stat[sheet.title][ps.NOTHING][kk[kaf]] += int(not fos and not test)

        print(i, kaf, fos, test)

sheet = wb.create_sheet('stat')

i = 1
for plan in stat:

    sheet[f"A{i}"].value = plan
    i += 1

    sheet[f"A{i}"].value = "Статистика"
    sheet[f"B{i}"].value = "Все"
    sheet[f"C{i}"].value = kk[1]
    sheet[f"D{i}"].value = kk[0]
    i += 1

    for row in stat[plan]:

        sheet[f"A{i}"].value = row
        sheet[f"B{i}"].value = stat[plan][row]['Все']
        sheet[f"C{i}"].value = stat[plan][row][kk[1]]
        sheet[f"D{i}"].value = stat[plan][row][kk[0]]
        i += 1

    i+=1

wb.save('./dest/test.xlsx')
